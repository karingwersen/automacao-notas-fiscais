
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from time import sleep
from shutil import move


def processar_planilha():

    path = "C:\\Users\\karin\\Downloads\\PLANILHA NF'S 2018.xlsx"
    planilha = openpyxl.load_workbook(path)
    sheet = planilha.active

    max_row = sheet.max_row
    for i in range(1, max_row + 1):
        cell = sheet.cell(row = i, column = 1)
        print(cell.value)


def acessar_nota(chave:str):

    chrome_options = Options()
    chrome_options.add_extension("fsist.crx")

    driver = webdriver.Chrome(options=chrome_options)
    driver.get("https://www.fsist.com.br/")

    caixa_texto = driver.find_element(By.ID, "chave")
    caixa_texto.send_keys(chave)
    consulta_nota = driver.find_element(By.ID, "butconsulta")
    consulta_nota.click()

    sleep(30)
    driver.quit()


def mover_arquivo(caminho_origem:str, caminho_destino:str):

    move(caminho_origem, caminho_destino)


if __name__ == '__main__':



    pass